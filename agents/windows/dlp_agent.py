# -*- coding: utf-8 -*-
import os, re, json, time, requests, win32clipboard
from datetime import datetime
from pathlib import Path
from threading import Thread
from openpyxl import load_workbook
from docx import Document
from PyPDF2 import PdfReader

# ---------------- CONFIG ----------------
CONFIG = {
    "scan_dirs": [
        os.path.expanduser("~/Documents"),
        os.path.expanduser("~/Desktop")
    ],
    "file_extensions": [".txt", ".csv", ".xlsx", ".xls", ".docx", ".pdf"],
    "server_url": "http://127.0.0.1:8000/api/events",
    "jwt": "secrettoken",
    "patterns": {
        
        "credit_cards": r"\b(?:\d[ -]*?){13,16}\b",
        "api_keys": r"(?:api|secret|token|key)[-_]?[\w]{16,40}",
        "chatgpt_domains": r"(chatgpt\.com|openai\.com|claude\.ai|bard\.google\.com)",
        "classified": r"(confidential|secret|restricted|internal use only)"
    }
}

PATTERNS = {k: re.compile(v, re.IGNORECASE) for k, v in CONFIG["patterns"].items()}

# ---------------- UTILITIES ----------------
def detect_sensitive(text):
    hits = {}
    for name, pattern in PATTERNS.items():
        matches = pattern.findall(text)
        if matches:
            hits[name] = matches
    return hits

def report_event(device_id, user_email, event_type, target, snippet, detector_hits):
    payload = {
        "device_id": device_id,
        "user_email": user_email,
        "event_type": event_type,
        "target": target,
        "snippet": snippet[:200],
        "detector_hits": detector_hits,
        "policy_id": None
    }
    headers = {"Authorization": f"Bearer {CONFIG['jwt']}"}
    try:
        requests.post(CONFIG["server_url"], json=payload, headers=headers, timeout=3)
    except Exception as e:
        print(f"[!] Reporting failed: {e}")

# ---------------- FILE SCANNING ----------------
def scan_txt_csv(filepath):
    try:
        with open(filepath, "r", errors="ignore") as f:
            text = f.read()
        return detect_sensitive(text)
    except:
        return {}

def scan_docx(filepath):
    try:
        doc = Document(filepath)
        full_text = "\n".join([p.text for p in doc.paragraphs])
        return detect_sensitive(full_text)
    except:
        return {}

def scan_xlsx(filepath):
    hits = {}
    try:
        wb = load_workbook(filepath, data_only=True)
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cell_hits = detect_sensitive(str(cell))
                        for k,v in cell_hits.items():
                            hits.setdefault(k, []).extend(v)
    except:
        pass
    return hits

def scan_pdf(filepath):
    hits = {}
    try:
        reader = PdfReader(filepath)
        for page in reader.pages:
            text = page.extract_text() or ""
            page_hits = detect_sensitive(text)
            for k,v in page_hits.items():
                hits.setdefault(k, []).extend(v)
    except:
        pass
    return hits

def scan_file(filepath):
    ext = filepath.suffix.lower()
    if ext in [".txt", ".csv"]:
        return scan_txt_csv(filepath)
    elif ext == ".docx":
        return scan_docx(filepath)
    elif ext in [".xlsx", ".xls"]:
        return scan_xlsx(filepath)
    elif ext == ".pdf":
        return scan_pdf(filepath)
    return {}

def scan_dirs():
    device_id = os.environ.get("COMPUTERNAME", "local_device")
    user_email = os.getlogin()
    total_files = 0
    for base_dir in CONFIG["scan_dirs"]:
        base_path = Path(base_dir)
        for file in base_path.rglob("*"):
            if file.suffix.lower() in CONFIG["file_extensions"]:
                total_files += 1
                hits = scan_file(file)
                if hits:
                    print(f"[+] Sensitive data found in {file}: {hits}")
                    report_event(device_id, user_email, "file_scan", str(file), str(file), hits)
            if total_files % 50 == 0:
                print(f"[*] Scanned {total_files} files...")

# ---------------- CLIPBOARD MONITOR ----------------
def monitor_clipboard():
    last_data = ""
    device_id = os.environ.get("COMPUTERNAME", "local_device")
    user_email = os.getlogin()
    while True:
        try:
            win32clipboard.OpenClipboard()
            data = win32clipboard.GetClipboardData()
            win32clipboard.CloseClipboard()
        except:
            data = ""
        if data and data != last_data:
            hits = detect_sensitive(data)
            if hits:
                print(f"[+] Clipboard sensitive data blocked: {hits} -> {data}")
                report_event(device_id, user_email, "clipboard", "clipboard", data, hits)
                # clear clipboard
                try:
                    win32clipboard.OpenClipboard()
                    win32clipboard.EmptyClipboard()
                    win32clipboard.CloseClipboard()
                except:
                    pass
            last_data = data
        time.sleep(1)

# ---------------- MAIN ----------------
if __name__ == "__main__":
    # Start clipboard monitoring
    Thread(target=monitor_clipboard, daemon=True).start()

    # Start file scanning
    def scan_worker():
        print("[*] Starting local sensitive data scan...")
        scan_dirs()
        print("[*] Scan complete.")

    Thread(target=scan_worker, daemon=True).start()

    print("[*] Agent running. Clipboard monitor and file scanner active...")
    while True:
        time.sleep(5)
