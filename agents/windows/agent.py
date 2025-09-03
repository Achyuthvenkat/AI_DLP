# agent.py

import os, re, time, logging, requests, threading, json
from pathlib import Path
import win32clipboard
from openpyxl import load_workbook
from docx import Document
from PyPDF2 import PdfReader
from dotenv import load_dotenv
from openai import OpenAI
from pynput import keyboard
import pickle
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
import random
from config import CONFIG, EXCLUDE_DIRS, OUTPUT_SCHEMA

# ---------------- ENV & CONFIG ----------------
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
SERVER_URL = os.getenv("SERVER_URL")
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

DEBUG_CONSOLE = True

# Add all drives except excluded
for drive_letter in range(65, 91):
    drive = f"{chr(drive_letter)}:\\"
    if os.path.exists(drive):
        CONFIG["scan_dirs"].append(drive)

PATTERNS = {k: re.compile(v, re.IGNORECASE) for k, v in CONFIG["patterns"].items()}

# ---------------- LOGGING ----------------
logging.basicConfig(
    filename="agent.log",
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# ---------------- GLOBAL STATS ----------------
stats = {"files_scanned": 0, "hits_detected": 0, "reports_sent": 0}

# ---------------- FILE TRACKING ----------------
existing_files_in_db = set()
scanned_files = {}
findings_summary = []
findings_lock = threading.Lock()

# ---------------- ML MODEL ----------------
sklearn_model = None
tfidf_vectorizer = None


def verhoeff_check(num_string):
    """Verhoeff algorithm check for Aadhaar numbers"""

    def d(j, k):
        multiplication_table = [
            [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
            [1, 2, 3, 4, 0, 6, 7, 8, 9, 5],
            [2, 3, 4, 0, 1, 7, 8, 9, 5, 6],
            [3, 4, 0, 1, 2, 8, 9, 5, 6, 7],
            [4, 0, 1, 2, 3, 9, 5, 6, 7, 8],
            [5, 9, 8, 7, 6, 0, 4, 3, 2, 1],
            [6, 5, 9, 8, 7, 1, 0, 4, 3, 2],
            [7, 6, 5, 9, 8, 2, 1, 0, 4, 3],
            [8, 7, 6, 5, 9, 3, 2, 1, 0, 4],
            [9, 8, 7, 6, 5, 4, 3, 2, 1, 0],
        ]
        return multiplication_table[j][k]

    def p(i, ni):
        permutation_table = [
            [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
            [1, 5, 7, 6, 2, 8, 3, 0, 9, 4],
            [5, 8, 0, 3, 7, 9, 6, 1, 4, 2],
            [8, 9, 1, 6, 0, 4, 3, 5, 2, 7],
            [9, 4, 5, 3, 1, 2, 6, 8, 7, 0],
            [4, 2, 8, 6, 5, 7, 3, 9, 0, 1],
            [2, 7, 9, 3, 8, 0, 6, 4, 1, 5],
            [7, 0, 4, 6, 9, 1, 3, 2, 5, 8],
        ]
        return permutation_table[i % 8][ni]

    try:
        c = 0
        for i, n in enumerate(reversed([int(x) for x in num_string])):
            c = d(c, p(i, n))
        return c == 0
    except:
        return False


def luhn_check(card_number):
    """Luhn algorithm check for credit card numbers"""
    try:
        digits = [int(x) for x in str(card_number).replace(" ", "").replace("-", "")]
        for i in range(len(digits) - 2, -1, -2):
            digits[i] *= 2
            if digits[i] > 9:
                digits[i] -= 9
        return sum(digits) % 10 == 0
    except:
        return False


# For Scikit learn model training
def create_training_data():
    """Create enhanced synthetic training data for scikit-learn model with proper confidence mapping"""
    training_data = []

    # Confidential examples (0.75-1.0 confidence range)
    confidential_samples = [
        # API Keys and Tokens
        (
            "OpenAI API key: sk-abc123xyz456def789ghi012jkl345mno678pqr901stu234",
            "Confidential",
        ),
        ("Gemini API key: AIzaSyDaGmWKa4JsXZ-HjGw7ISLn_3namBGewQe", "Confidential"),
        ("DeepSeek API: ds-abc123xyz456def789", "Confidential"),
        (
            "AWS_SECRET_ACCESS_KEY=wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
            "Confidential",
        ),
        ("AZURE_CLIENT_SECRET: abc123def456ghi789", "Confidential"),
        ("Bearer token: eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9", "Confidential"),
        ("JWT TOKEN: eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9", "Confidential"),
        # Database Credentials
        ("Database password: db_pass=SecretP@ssw0rd123", "Confidential"),
        ("MYSQL_ROOT_PASSWORD=SuperSecret123!", "Confidential"),
        ("postgresql://user:password123@localhost:5432/database", "Confidential"),
        ("mongodb://admin:secret123@cluster.mongodb.net/", "Confidential"),
        # Private Keys and Certificates
        ("-----BEGIN RSA PRIVATE KEY-----", "Confidential"),
        ("-----BEGIN PRIVATE KEY-----", "Confidential"),
        ("-----BEGIN ENCRYPTED PRIVATE KEY-----", "Confidential"),
        ("-----BEGIN OPENSSH PRIVATE KEY-----", "Confidential"),
        # Passwords and Authentication
        ('password="MySecretPassword123"', "Confidential"),
        ("admin_password: SuperSecure@2024", "Confidential"),
        ("root:$6$salt$hashedpassword", "Confidential"),
        ("SECRET_KEY = 'django-insecure-abcd1234'", "Confidential"),
        # Company Confidential Information
        ("Titan Company Limited R&D blueprint for Orion project", "Confidential"),
        ("CONFIDENTIAL Titan financial report Q3-2024", "Confidential"),
        ("TOP SECRET merger acquisition documents", "Confidential"),
        ("Confidential source code repository access", "Confidential"),
        ("Trade secret formula for new product development", "Confidential"),
        ("Legal privileged attorney-client communication", "Confidential"),
        # Security and Encryption
        ("AES encryption key: 256bit-abcd1234efgh5678", "Confidential"),
        ("SSL certificate private key", "Confidential"),
        ("Two-factor authentication backup codes", "Confidential"),
        ("Security vulnerability assessment report", "Confidential"),
    ]

    # Sensitive examples (0.50-0.75 confidence range)
    sensitive_samples = [
        # Personal Identification Numbers
        ("PAN number: FURPA3446B", "Sensitive"),
        ("PAN: ABCDE1234F", "Sensitive"),
        ("Aadhaar number: 123456789012", "Sensitive"),
        ("Aadhaar: 1234 5678 9012", "Sensitive"),
        ("Social Security Number: 123-45-6789", "Sensitive"),
        ("Driver License: DL123456789", "Sensitive"),
        ("Passport Number: A12345678", "Sensitive"),
        ("Voter ID: ABC1234567", "Sensitive"),
        # Financial Information
        ("Credit card number: 4532-1234-5678-9012", "Sensitive"),
        ("Credit card: 4111111111111111", "Sensitive"),
        ("Bank account: 123456789012", "Sensitive"),
        ("IBAN: GB29 NWBK 6016 1331 9268 19", "Sensitive"),
        ("SWIFT code with account details", "Sensitive"),
        # Healthcare Information
        ("Medical record number: MR123456", "Sensitive"),
        ("Patient ID: PAT-2024-001", "Sensitive"),
        ("Health insurance policy: HI789012345", "Sensitive"),
        ("Medical diagnosis: diabetes mellitus type 2", "Sensitive"),
        ("Prescription details for patient treatment", "Sensitive"),
        ("Laboratory test results confidential", "Sensitive"),
        # Employment and HR Data
        ("Employee ID: EMP2024001", "Sensitive"),
        ("Salary information: Rs. 850,000 annual", "Sensitive"),
        ("HR document: Employee salary slip for March 2024", "Sensitive"),
        ("Performance review confidential rating", "Sensitive"),
        ("Employee background verification report", "Sensitive"),
        ("Payroll processing data for employees", "Sensitive"),
        ("Termination letter with severance details", "Sensitive"),
        # Contact and Personal Information
        ("Email: john.doe@company.com Phone: +91-9876543210", "Sensitive"),
        ("Address: 123 Main Street, Mumbai 400001", "Sensitive"),
        ("Date of birth: 15/08/1985", "Sensitive"),
        ("Emergency contact information", "Sensitive"),
    ]

    # Internal examples (0.25-0.50 confidence range)
    internal_samples = [
        # Meeting and Communication
        ("Meeting notes from quarterly business review", "Internal"),
        ("Internal email regarding project timeline updates", "Internal"),
        ("Conference call minutes for team sync", "Internal"),
        ("Action items from leadership meeting", "Internal"),
        ("Weekly status report for internal circulation", "Internal"),
        # Project and Operations
        ("Project kickoff scheduled for next month", "Internal"),
        ("Internal project timeline and milestones", "Internal"),
        ("Resource allocation for Q4 projects", "Internal"),
        ("Team capacity planning document", "Internal"),
        ("Project risk assessment internal review", "Internal"),
        # Financial and Budget
        ("Budget allocation for department Q4", "Internal"),
        ("Cost center analysis for internal review", "Internal"),
        ("Quarterly expense report summary", "Internal"),
        ("Internal audit findings report", "Internal"),
        ("Vendor contract negotiation notes", "Internal"),
        # Process and Documentation
        ("For internal use only - process documentation", "Internal"),
        ("Standard operating procedures update", "Internal"),
        ("Internal training materials draft", "Internal"),
        ("Quality assurance checklist internal", "Internal"),
        ("Change management process guidelines", "Internal"),
        # Strategy and Planning
        ("Strategic planning session outcomes", "Internal"),
        ("Market analysis for internal planning", "Internal"),
        ("Competitive intelligence briefing", "Internal"),
        ("Product roadmap discussion points", "Internal"),
        ("Internal feedback on new initiatives", "Internal"),
    ]

    # Public examples (0.0-0.25 confidence range)
    public_samples = [
        # Press Releases and Announcements
        (
            "Titan Company Limited announces launch of new jewellery collection on March 15",
            "Public",
        ),
        ("Press release: Titan Q2 sales grew by 12% year-over-year", "Public"),
        ("Company announces expansion into new markets", "Public"),
        ("Titan wins industry award for innovation", "Public"),
        ("New store opening celebration event", "Public"),
        # Marketing and Promotional
        ("Marketing flyer for Titan Watches – Flat 20% discount offer", "Public"),
        ("This brochure describes features of Titan's smart wearables", "Public"),
        ("Summer collection now available in stores", "Public"),
        ("Customer testimonials and success stories", "Public"),
        ("Product catalog available for download", "Public"),
        # Website and General Information
        ("Welcome to our official company website", "Public"),
        ("Titan website content: Careers page updated with new positions", "Public"),
        ("Contact us at support@titan.com for assistance", "Public"),
        ("Company history and milestones", "Public"),
        ("About us - company vision and mission", "Public"),
        ("Store locator and contact information", "Public"),
        # General Business Information
        ("Public financial statements and annual reports", "Public"),
        ("Industry news and market trends", "Public"),
        ("Company newsletter for customers", "Public"),
        ("Social media posts and updates", "Public"),
        ("Customer service FAQ section", "Public"),
        ("Terms of service and privacy policy", "Public"),
        # Events and Community
        ("Community outreach program details", "Public"),
        ("Corporate social responsibility initiatives", "Public"),
        ("Employee volunteer activities", "Public"),
        ("Industry conference participation", "Public"),
        ("Awards and recognition ceremonies", "Public"),
    ]

    # Combine all samples
    all_samples = (
        confidential_samples + sensitive_samples + internal_samples + public_samples
    )

    # Add base samples
    for text, label in all_samples:
        training_data.append((text, label))

    # Add variations with different formatting and context
    for text, label in all_samples:
        # Uppercase variations
        training_data.append((text.upper(), label))

        # Lowercase variations
        training_data.append((text.lower(), label))

        # Mixed case variations
        training_data.append((text.swapcase(), label))

        # Add context variations for more realistic scenarios
        if label == "Confidential":
            training_data.append((f"URGENT: {text}", label))
            training_data.append((f"Please handle this confidentially: {text}", label))
            training_data.append((f"Restricted access: {text}", label))

        elif label == "Sensitive":
            training_data.append((f"Personal information: {text}", label))
            training_data.append((f"Employee data: {text}", label))
            training_data.append((f"Customer record: {text}", label))

        elif label == "Internal":
            training_data.append((f"Internal memo: {text}", label))
            training_data.append((f"For team use: {text}", label))
            training_data.append((f"Department update: {text}", label))

        elif label == "Public":
            training_data.append((f"Public announcement: {text}", label))
            training_data.append((f"Customer information: {text}", label))
            training_data.append((f"Marketing material: {text}", label))

    # Add some edge cases and mixed content
    edge_cases = [
        ("Meeting about public product launch strategy", "Internal"),
        ("Internal discussion of public press release", "Internal"),
        ("Customer complaint containing personal details", "Sensitive"),
        ("Public API documentation with example keys", "Internal"),
        ("Marketing campaign targeting internal employees", "Internal"),
        ("Public statement about confidential investigation", "Public"),
        ("Customer service script for handling sensitive data", "Internal"),
        ("Training manual on handling confidential information", "Internal"),
        ("Public job posting for security role", "Public"),
        ("Internal security audit of public website", "Internal"),
    ]

    training_data.extend(edge_cases)

    return training_data


def train_sklearn_model():
    """Train and save the scikit-learn model"""
    global sklearn_model, tfidf_vectorizer

    try:
        # Create training data
        training_data = create_training_data()
        texts, labels = zip(*training_data)

        # Split the data
        X_train, X_test, y_train, y_test = train_test_split(
            texts, labels, test_size=0.2, random_state=42, stratify=labels
        )

        # Create pipeline
        sklearn_model = Pipeline(
            [
                (
                    "tfidf",
                    TfidfVectorizer(
                        max_features=1000, stop_words="english", ngram_range=(1, 2)
                    ),
                ),
                ("classifier", LogisticRegression(random_state=42, multi_class="ovr")),
            ]
        )

        # Train the model
        sklearn_model.fit(X_train, y_train)

        # Test the model
        y_pred = sklearn_model.predict(X_test)
        debug_print(f"[SKLEARN MODEL] Training completed")
        debug_print(
            f"[SKLEARN MODEL] Test accuracy: {sklearn_model.score(X_test, y_test):.3f}"
        )

        # Save the model
        with open(CONFIG["sklearn_classification"]["model_path"], "wb") as f:
            pickle.dump(sklearn_model, f)

        debug_print(
            f"[SKLEARN MODEL] Model saved to {CONFIG['sklearn_classification']['model_path']}"
        )

    except Exception as e:
        debug_print(f"[SKLEARN MODEL ERROR] {e}")
        logging.error(f"Sklearn model training failed: {e}")


def load_sklearn_model():
    """Load the scikit-learn model"""
    global sklearn_model

    model_path = CONFIG["sklearn_classification"]["model_path"]

    if os.path.exists(model_path):
        try:
            with open(model_path, "rb") as f:
                sklearn_model = pickle.load(f)
            debug_print(f"[SKLEARN MODEL] Loaded from {model_path}")
        except Exception as e:
            debug_print(f"[SKLEARN MODEL LOAD ERROR] {e}")
            train_sklearn_model()
    else:
        debug_print(f"[SKLEARN MODEL] Model not found, training new model...")
        train_sklearn_model()


def sklearn_classify(text: str):
    """Enhanced classification with proper confidence mapping to specified ranges"""
    if not CONFIG["sklearn_classification"]["enabled"] or not sklearn_model:
        return {"label": "N/A", "confidence": 0.0}

    try:
        # Get prediction and probability
        prediction = sklearn_model.predict([text])[0]
        probabilities = sklearn_model.predict_proba([text])[0]

        # Get the raw confidence (highest probability)
        raw_confidence = float(max(probabilities))

        # Map confidence to specified ranges based on classification
        if prediction == "Confidential":
            # Map to 0.75-1.0 range
            mapped_confidence = 0.75 + (raw_confidence * 0.25)
        elif prediction == "Sensitive":
            # Map to 0.50-0.75 range
            mapped_confidence = 0.50 + (raw_confidence * 0.25)
        elif prediction == "Internal":
            # Map to 0.25-0.50 range
            mapped_confidence = 0.25 + (raw_confidence * 0.25)
        elif prediction == "Public":
            # Map to 0.0-0.25 range
            mapped_confidence = 0.0 + (raw_confidence * 0.25)
        else:
            # Default to Internal range for unknown labels
            mapped_confidence = 0.25 + (raw_confidence * 0.25)

        # Ensure confidence is within bounds and rounded
        mapped_confidence = max(0.0, min(1.0, round(mapped_confidence, 3)))

        result = {"label": prediction, "confidence": mapped_confidence}
        debug_print(
            f"[SKLEARN CLASSIFICATION] Raw: {raw_confidence:.3f} -> Mapped: {mapped_confidence:.3f} ({prediction})"
        )

        return result

    except Exception as e:
        debug_print(f"[SKLEARN CLASSIFICATION ERROR] {e}")
        return {"label": "Internal", "confidence": 0.375}  # Middle of Internal range


# ---------------- UTILS ----------------
# Pre-compile regex patterns once
EXCLUDE_PATTERNS = []
for excl in EXCLUDE_DIRS:
    try:
        # Convert wildcard-style patterns to regex
        pattern = re.compile(
            excl.replace("\\", "\\\\").replace(".", r"\.").replace("*", ".*"),
            re.IGNORECASE,
        )
        EXCLUDE_PATTERNS.append(pattern)
    except:
        continue


def should_exclude(path: str) -> bool:
    """
    Return True if the path matches any exclusion rule.
    Supports absolute paths, folder names, and regex patterns.
    """
    path_obj = Path(path).resolve()
    path_str = str(path_obj).replace("/", "\\")
    path_lower = path_str.lower()

    for pattern in EXCLUDE_PATTERNS:
        if pattern.search(path_lower):
            return True

    # Folder name check
    for part in path_obj.parts:
        if part.lower() in [ex.lower() for ex in EXCLUDE_DIRS]:
            return True

    return False


def debug_print(msg):
    if DEBUG_CONSOLE:
        print(msg)
    logging.info(msg)


def get_file_hash(filepath: Path):
    """Simple file hash based on size and modification time"""
    try:
        stat = filepath.stat()
        return f"{stat.st_size}_{int(stat.st_mtime)}"
    except:
        return None


# ---------------- JWT FETCH ----------------
def fetch_jwt_token():
    device_id = os.environ.get("COMPUTERNAME", "local_device")
    try:
        resp = requests.post(
            f"{SERVER_URL}/api/token", json={"device_id": device_id}, timeout=5
        )
        if resp.status_code == 200:
            response_data = resp.json()
            token = response_data.get("token")
            existing_files = response_data.get("existing_files", [])

            global existing_files_in_db
            existing_files_in_db = set(existing_files)

            if token:
                debug_print(
                    f"[JWT FETCHED] {token[:8]}... with {len(existing_files)} existing files"
                )
                return token
        debug_print(f"[JWT FETCH FAILED] {resp.status_code}: {resp.text}")
        return None
    except Exception as e:
        debug_print(f"[JWT FETCH ERROR] {e}")
        return None


def sync_file_states():
    """Sync current file state with server"""
    device_id = os.environ.get("COMPUTERNAME", "local_device")

    current_files = []
    for base_dir in CONFIG["scan_dirs"]:
        if should_exclude(base_dir):
            continue
        try:
            for file in Path(base_dir).rglob("*"):
                if (
                    file.is_file()
                    and file.suffix.lower() in CONFIG["file_extensions"]
                    and not should_exclude(str(file))
                ):
                    current_files.append(str(file))
        except Exception as e:
            logging.error(f"Error scanning {base_dir}: {e}")

    try:
        headers = {"Authorization": f"Bearer {JWT_TOKEN}"}
        resp = requests.post(
            CONFIG["sync_url"],
            json={"device_id": device_id, "current_files": current_files},
            headers=headers,
            timeout=10,
        )

        if resp.status_code == 200:
            result = resp.json()
            debug_print(
                f"[FILE SYNC] Deleted: {result['deleted_files_count']}, New files to scan: {len(result['new_files_to_scan'])}"
            )
            return result["new_files_to_scan"]
        else:
            debug_print(f"[FILE SYNC FAILED] {resp.status_code}: {resp.text}")
            return current_files

    except Exception as e:
        debug_print(f"[FILE SYNC ERROR] {e}")
        return current_files


JWT_TOKEN = fetch_jwt_token() or "secrettoken"


# ---------------- ENHANCED DETECTION ----------------
def detect_sensitive(text: str, strict_validation=False):
    hits = {}
    if not text:
        return hits

    for name, pattern in PATTERNS.items():
        found = pattern.findall(text)
        if found:
            if strict_validation:
                # Additional validation for specific patterns
                if name == "aadhaar_strict":
                    # Validate with Verhoeff algorithm
                    valid_aadhaar = [
                        num for num in found if len(num) == 12 and verhoeff_check(num)
                    ]
                    if valid_aadhaar:
                        hits[name] = valid_aadhaar
                elif name == "credit_card_strict":
                    # Validate with Luhn algorithm
                    valid_cards = []
                    for card in found:
                        clean_card = re.sub(r"[- ]", "", card)
                        if 15 <= len(clean_card) <= 19 and luhn_check(clean_card):
                            valid_cards.append(card)
                    if valid_cards:
                        hits[name] = valid_cards
                elif name == "pan_strict":
                    # Validate PAN format strictly
                    valid_pans = [
                        pan for pan in found if len(pan) == 10 and pan.isalnum()
                    ]
                    if valid_pans:
                        hits[name] = valid_pans
            else:
                hits[name] = found

    return hits


# ---------------- AI CLASSIFICATION ----------------
def ai_classify(text: str):
    if not CONFIG["ai_classification"]["enabled"] or not client:
        return {"label": "N/A", "confidence": 0.0}

    try:
        response = client.responses.create(
            instructions="Give a valid reason in output for the classification.",
            model=CONFIG["ai_classification"]["model"],
            input=[
                {
                    "role": "developer",
                    "content": (
                        "You are a Data Loss Prevention (DLP) classifier. Analyze the following text and classify it into one of these security categories.\n\n"
                        "CATEGORIES & CONFIDENCE RANGES:\n"
                        "If you recieve a filepath or file name instead of content it means it has no sensitive content, so classify it accordingly.\n\n"
                        "• Public (0.0-0.25): Marketing content, press releases, website information, product brochures\n"
                        "• Internal (0.25-0.50): Meeting notes, internal emails, project timelines, internal documentation, budget allocations\n"
                        "• Sensitive (0.50-0.75): Personally Identifiable Information (PII) such as PAN, Aadhaar, credit card numbers, medical reports, salary slips, employee payroll information\n"
                        "• Confidential (0.75-1.0): API keys, passwords, tokens, private keys, database credentials, trade secrets, legal documents, Titan R&D blueprints, financial reports, source code, security credentials\n\n"
                        "EXAMPLES:\n"
                        "Confidential:\n"
                        "- OpenAI API key: sk-abc123...\n"
                        "- Gemini API key: AIzaSy...\n"
                        "- DeepSeek API: ds-abc123...\n"
                        "- Database password: db_pass=SecretP@ssw0rd123\n"
                        "- -----BEGIN RSA PRIVATE KEY-----\n"
                        "- Titan Company Limited R&D blueprint for Orion project\n"
                        "- Confidential Titan financial report Q3-2024\n"
                        '- password="MySecretPassword123"\n\n'
                        "Sensitive:\n"
                        "- PAN number: FURPA3446B \n"
                        "- PAN regex: [A-Za-z]{5}[0-9]{4}[A-Za-z]{1}\n"
                        "- Aadhaar: 123456789012\n"
                        "- Aadhaar regex: \d{4}[\s-]?\d{4}[\s-]?\d{4}\n"
                        "- Credit card: 4532-1234-5678-9012\n"
                        "- Employee salary slip for March 2024\n"
                        "- Medical report showing diagnosis\n"
                        "- Employee payroll information\n\n"
                        "Internal:\n"
                        "- Meeting notes from quarterly review\n"
                        "- Internal email regarding project timeline\n"
                        "- Team budget allocation for Q4\n"
                        "- Project kickoff schedule\n"
                        "- Process documentation (internal use only)\n"
                        "- Minutes of meeting - action items\n\n"
                        "Public:\n"
                        "- Titan Company Limited announces jewellery collection launch\n"
                        "- Brochure describing Titan smart wearables\n"
                        "- Press release: Titan Q2 sales grew by 12%\n"
                        "- Marketing flyer for Titan Watches – 20% discount\n"
                        "- Titan website content (e.g., careers page)\n"
                        "- Official website welcome text\n"
                        "- Public product catalog\n\n"
                        """ ### Indicators:
                        Confidential → API keys, passwords, tokens, private keys, secret config  
                        Sensitive → Aadhaar number, PAN number, SSN, credit card numbers, medical data  
                        Internal → Business documents, financial reports, HR info, project data not public  
                        Public → Marketing material, press releases, general info, public website text  

                        ### Examples:
                        Example: "OpenAI API key: sk-abc123xyz456" → Confidential  
                        Example: "Password=Titan@123" → Confidential  
                        Example: "PAN number: ABCDE1234F" → Sensitive  
                        Example: "Aadhaar: 1234 5678 9012" → Sensitive  
                        Example: "Quarterly Financial Report 2023" → Internal  
                        Example: "Titan launches new smartwatch" → Public  
                        Example: "Our internal project code name is Project-X" → Internal  
                        Example: "Contact us at support@titan.com" → Public  

                        Now classify the following text:"""
                        "RESPONSE FORMAT:\n"
                        "Respond with ONLY valid JSON in this exact format:\n"
                        '{"label": "Confidential", "confidence": 0.92}\n'
                    ),
                },
                {
                    "role": "user",
                    "content": text,
                },
            ],
            text={"format": OUTPUT_SCHEMA},
            temperature=0.3,
        )
        print(response)

        raw_content = (response.output_text or "").strip()
        debug_print(f"[AI RAW RESPONSE] {raw_content}")
        parsed = json.loads(raw_content)

        label = parsed.get("label", "Unknown")
        confidence = parsed.get("confidence", 0.0)

        try:
            confidence = float(confidence)
            confidence = max(0.0, min(1.0, round(confidence, 3)))
        except (ValueError, TypeError):
            confidence = 0.0

        valid_labels = ["Public", "Internal", "Sensitive", "Confidential", "safe"]
        if label not in valid_labels:
            debug_print(
                f"[AI LABEL ERROR] Invalid label '{label}', defaulting to 'Internal'"
            )
            label = "safe"
            confidence = 0.0

        result = {"label": label, "confidence": confidence}
        debug_print(f"[AI CLASSIFICATION SUCCESS] {result}")
        return result

    except json.JSONDecodeError as e:
        debug_print(f"[AI JSON PARSE ERROR] {e}")
        return {"label": "Internal", "confidence": 0.4}
    except Exception as e:
        print(f"[AI CLASSIFICATION ERROR] {e}")

        return {"label": "Internal", "confidence": 0.4}


# ---------------- SUMMARY BUFFER ----------------
def add_to_summary(event_type, target, snippet, hits):
    # Get both AI and sklearn classifications
    ai_result = ai_classify(snippet)
    sklearn_result = sklearn_classify(snippet)

    with findings_lock:
        findings_summary.append(
            {
                "device_id": os.environ.get("COMPUTERNAME", "local_device"),
                "user_email": os.getlogin(),
                "event_type": event_type,
                "target": str(target),
                "snippet": (snippet or "")[:200],
                "detector_hits": hits,
                "ai_classification": ai_result,
                "sklearn_classification": sklearn_result,
            }
        )
    stats["hits_detected"] += len(hits)


def send_summary_to_server():
    while True:
        time.sleep(60)
        to_send = None
        with findings_lock:
            if findings_summary:
                to_send = list(findings_summary)
                findings_summary.clear()
        if to_send:
            headers = {"Authorization": f"Bearer {JWT_TOKEN}"}
            try:
                response = requests.post(
                    CONFIG["server_url"],
                    json={"events": to_send},
                    headers=headers,
                    timeout=20,
                )
                if response.status_code == 200:
                    debug_print(f"[SUMMARY SENT] {len(to_send)} findings")
                    stats["reports_sent"] += 1
                else:
                    debug_print(
                        f"[SUMMARY FAILED] {response.status_code}: {response.text}"
                    )
                    with findings_lock:
                        findings_summary[:0] = to_send
            except Exception as e:
                logging.error(f"Failed to send summary: {e}")
                with findings_lock:
                    findings_summary[:0] = to_send


# ---------------- FILE SCANNING ----------------
def scan_file(filepath: Path, force_scan=False):
    file_str = str(filepath)
    file_hash = get_file_hash(filepath)

    if not force_scan and file_str in scanned_files:
        if scanned_files[file_str] == file_hash:
            return

    debug_print(f"[SCANNING] {filepath}")

    hits = {}
    sensitive_context = []  # Store context around sensitive data
    ext = filepath.suffix.lower()

    try:
        if ext in [".txt", ".csv", ".py", ".js", ".java", ".go", ".ts"]:
            with open(filepath, "r", errors="ignore") as f:
                lines = f.readlines()

            for idx, line in enumerate(lines, start=1):
                line_hits = detect_sensitive(line)
                if line_hits:
                    # Get context: current line + surrounding lines
                    context_lines = []
                    for i in range(max(0, idx - 2), min(len(lines), idx + 1)):
                        context_lines.append(f"L{i+1}: {lines[i].strip()}")

                    context = "\n".join(context_lines)
                    sensitive_context.append(context[:300])

                    for k, v in line_hits.items():
                        hits.setdefault(k, []).append(
                            {"line": idx, "snippet": line[:200], "matches": v}
                        )

        elif ext == ".docx":
            doc = Document(filepath)
            paragraphs = [p.text for p in doc.paragraphs]

            for idx, p_text in enumerate(paragraphs, start=1):
                line_hits = detect_sensitive(p_text)
                if line_hits:
                    # Get context: current paragraph + surrounding
                    context_paras = []
                    for i in range(max(0, idx - 2), min(len(paragraphs), idx + 1)):
                        if paragraphs[i].strip():  # Skip empty paragraphs
                            context_paras.append(f"P{i+1}: {paragraphs[i].strip()}")

                    context = "\n".join(context_paras)
                    sensitive_context.append(context[:300])

                    for k, v in line_hits.items():
                        hits.setdefault(k, []).append(
                            {"line": idx, "snippet": p_text[:200], "matches": v}
                        )

        elif ext in [".xlsx", ".xls"]:
            wb = load_workbook(filepath, data_only=True)
            for sheet in wb:
                for row_idx, row in enumerate(
                    sheet.iter_rows(values_only=True), start=1
                ):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell:
                            cell_hits = detect_sensitive(str(cell))
                            if cell_hits:
                                # Get cell context (nearby cells)
                                context_cells = []
                                for r in range(
                                    max(1, row_idx - 1),
                                    min(sheet.max_row + 1, row_idx + 2),
                                ):
                                    for c in range(
                                        max(1, col_idx - 1),
                                        min(sheet.max_column + 1, col_idx + 2),
                                    ):
                                        try:
                                            cell_val = sheet.cell(row=r, column=c).value
                                            if cell_val:
                                                context_cells.append(
                                                    f"R{r}C{c}: {str(cell_val)[:50]}"
                                                )
                                        except:
                                            pass

                                context = " | ".join(context_cells[:5])  # Limit context
                                sensitive_context.append(context[:300])

                                for k, v in cell_hits.items():
                                    hits.setdefault(k, []).append(
                                        {
                                            "line": f"Sheet:{sheet.title} Row:{row_idx} Col:{col_idx}",
                                            "snippet": str(cell)[:200],
                                            "matches": v,
                                        }
                                    )

        elif ext == ".pdf":
            reader = PdfReader(filepath)
            for page_idx, page in enumerate(reader.pages, start=1):
                text = page.extract_text() or ""
                lines = text.splitlines()

                for line_idx, line in enumerate(lines, start=1):
                    line_hits = detect_sensitive(line)
                    if line_hits:
                        # Get context: surrounding lines on the page
                        context_lines = []
                        for i in range(
                            max(0, line_idx - 2), min(len(lines), line_idx + 1)
                        ):
                            context_lines.append(f"L{i+1}: {lines[i].strip()}")

                        context = f"Page {page_idx}: " + "\n".join(context_lines)
                        sensitive_context.append(context[:300])

                        for k, v in line_hits.items():
                            hits.setdefault(k, []).append(
                                {
                                    "line": f"Page:{page_idx} Line:{line_idx}",
                                    "snippet": line[:200],
                                    "matches": v,
                                }
                            )

    except Exception as e:
        logging.error(f"Failed to scan {filepath}: {e}")
        return

    scanned_files[file_str] = file_hash
    stats["files_scanned"] += 1

    # Prepare enriched snippet for summary
    if sensitive_context:
        # Join contexts with clear separators
        combined_snippet = "\n--- CONTEXT ---\n".join(sensitive_context)
        if len(combined_snippet) > 800:  # Limit total length
            combined_snippet = combined_snippet[:797] + "..."

        # Add file metadata
        file_info = f"FILE: {filepath.name} ({ext.upper()}) | "
        combined_snippet = file_info + combined_snippet
    else:
        with open("errors.txt", "a", encoding="utf-8") as error_file:
            error_file.write(f"No sensitive content in {filepath}\n")
        combined_snippet = (
            f"FILE: {filepath.name} | No sensitive content context available"
        )

    if hits or file_str not in existing_files_in_db:
        if hits:
            debug_print(f"[FILE DETECTED] {filepath}")
            debug_print(f"[ENRICHED SNIPPET] {combined_snippet[:150]}...")

        # Pass the enriched snippet with context
        add_to_summary("file_scan", filepath, combined_snippet, hits)


def get_current_files():
    """
    Return all files to scan, skipping excluded directories.
    """
    current_files = []

    for base_dir in CONFIG["scan_dirs"]:
        base_path = Path(base_dir)
        if should_exclude(base_path):
            continue

        try:
            for root, dirs, files in os.walk(base_path):
                # Remove excluded dirs so os.walk doesn't descend into them
                dirs[:] = [d for d in dirs if not should_exclude(Path(root) / d)]

                for f in files:
                    file_path = Path(root) / f
                    if file_path.suffix.lower() in CONFIG["file_extensions"]:
                        if not should_exclude(file_path):
                            current_files.append(str(file_path))
        except Exception as e:
            logging.error(f"Error scanning {base_dir}: {e}")

    return current_files


def incremental_file_scan():
    debug_print("[INCREMENTAL SCAN] Starting file sync...")
    files_to_scan = sync_file_states()

    if not files_to_scan:
        debug_print("[INCREMENTAL SCAN] No new files to scan")
        return

    debug_print(f"[INCREMENTAL SCAN] Scanning {len(files_to_scan)} new files...")

    for file_path in files_to_scan:
        try:
            filepath = Path(file_path)
            if filepath.exists() and not should_exclude(filepath):
                scan_file(filepath, force_scan=True)
        except Exception as e:
            logging.error(f"Error scanning {file_path}: {e}")

    debug_print(f"[INCREMENTAL SCAN] Completed scanning {len(files_to_scan)} files")


def scan_dirs():
    incremental_file_scan()
    while True:
        time.sleep(1800)
        try:
            incremental_file_scan()
        except Exception as e:
            logging.error(f"Incremental scan error: {e}")


# ---------------- CLIPBOARD ----------------
def monitor_clipboard():
    last_data = ""
    while True:
        try:
            win32clipboard.OpenClipboard()
            data = win32clipboard.GetClipboardData()
            win32clipboard.CloseClipboard()
        except Exception:
            data = ""
        if isinstance(data, str) and data and data != last_data:
            hits = detect_sensitive(data)
            if hits:
                debug_print(f"[CLIPBOARD DETECTED] {hits}")
                add_to_summary("clipboard", "clipboard", data, hits)
            last_data = data
        time.sleep(1)


# ---------------- KEYLOGGER ----------------
typed_buffer = ""


def on_press(key):
    global typed_buffer
    try:
        if hasattr(key, "char") and key.char:
            typed_buffer += key.char
        elif key == keyboard.Key.space:
            typed_buffer += " "
        elif key == keyboard.Key.enter:
            typed_buffer += "\n"
        if len(typed_buffer) > 50 or "\n" in typed_buffer:
            hits = detect_sensitive(typed_buffer)
            if hits:
                debug_print(f"[KEYSTROKE DETECTED] {hits}")
                add_to_summary("keystroke", "keyboard", typed_buffer, hits)
            typed_buffer = ""
    except Exception as e:
        logging.error(f"Keylogger error: {e}")


def start_keylogger():
    listener = keyboard.Listener(on_press=on_press)
    listener.daemon = True
    listener.start()


# ---------------- MAIN ----------------
if __name__ == "__main__":
    debug_print("🚀 DLP Agent starting with dual AI+Sklearn classification...")

    # Load or train sklearn model
    load_sklearn_model()

    threading.Thread(target=monitor_clipboard, daemon=True).start()
    start_keylogger()
    threading.Thread(target=scan_dirs, daemon=True).start()
    threading.Thread(target=send_summary_to_server, daemon=True).start()
    debug_print(
        "📋 Clipboard, ⌨️ Keystrokes, and 📂 Incremental file scanning running..."
    )

    debug_print(
        f"[INITIAL STATE] {len(existing_files_in_db)} files already in database"
    )

    while True:
        time.sleep(5)
        if stats["files_scanned"] % 100 == 0 and stats["files_scanned"] > 0:
            debug_print(
                f"[STATS] Files: {stats['files_scanned']}, Hits: {stats['hits_detected']}, Reports: {stats['reports_sent']}"
            )
